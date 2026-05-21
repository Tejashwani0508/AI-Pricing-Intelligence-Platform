"""
Automated Reporting Module

Generates professional enterprise reports in three formats:
  - CSV  (portable data export)
  - Excel (formatted multi-sheet workbook with OpenPyXL)
  - PDF  (styled document with ReportLab)

Each report includes:
  - Executive KPIs (revenue, profit, margin, risk)
  - Portfolio summary (category breakdowns)
  - Risk analysis (score distribution, high-risk products)
  - Pricing recommendations (optimization summary)
  - Top/bottom product tables
  - Charts and visual elements

All functions accept a DataFrame and return the file path to the generated report.
"""

import io
import logging
import os
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

# ─── PDF (ReportLab) ───────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm, cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
    HRFlowable,
    KeepTogether,
    Frame,
    PageTemplate,
    BaseDocTemplate,
)

# ─── Matplotlib for chart generation in PDF ──────────────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import FancyBboxPatch

# ─── Excel (OpenPyXL) ─────────────────────────────────────────────────────
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle,
    numbers,
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from utils.config import AppConfig
from utils.helpers import format_currency, safe_divide

logger = logging.getLogger("ai_pricing.reporting")

# Suppress matplotlib font warnings
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")


# ═══════════════════════════════════════════════════════════════════════════
# CONSTANTS & STYLES
# ═══════════════════════════════════════════════════════════════════════════

# Brand colours
BRAND_PRIMARY = "#1a237e"
BRAND_SECONDARY = "#0d47a1"
BRAND_ACCENT = "#2e7d32"
BRAND_WARNING = "#f9a825"
BRAND_DANGER = "#c62828"
BRAND_LIGHT = "#f5f5f5"
BRAND_WHITE = "#ffffff"
BRAND_DARK = "#1a1a2e"
BRAND_TEAL = "#00838f"
BRAND_PURPLE = "#4a148c"
BRAND_GREY = "#757575"
BRAND_BG_LIGHT = "#fafafa"

# PDF page dimensions
PAGE_WIDTH, PAGE_HEIGHT = letter

# Chart colour palette for matplotlib (professional, presentation-friendly)
CHART_COLORS = ["#1a237e", "#0d47a1", "#2e7d32", "#f9a825", "#c62828",
                "#00838f", "#4a148c", "#e65100", "#37474f", "#6a1b9a"]
CHART_COLORS_SEQ = ["#e3f2fd", "#bbdefb", "#90caf9", "#64b5f6", "#42a5f5",
                    "#2196f3", "#1e88e5", "#1976d2", "#1565c0", "#0d47a1"]

RISK_LEVEL_ORDER = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]

# Column mapping for standardised report columns
DEFAULT_KPI_METRICS: List[Dict[str, str]] = [
    {"label": "Total Products", "key": "total_products", "prefix": "", "suffix": ""},
    {"label": "Total Revenue", "key": "total_revenue", "prefix": "$", "suffix": ""},
    {"label": "Total Profit", "key": "total_profit", "prefix": "$", "suffix": ""},
    {"label": "Avg Margin", "key": "avg_margin", "prefix": "", "suffix": "%"},
    {"label": "Avg Risk", "key": "avg_risk", "prefix": "", "suffix": ""},
    {"label": "High Risk", "key": "high_risk_count", "prefix": "", "suffix": ""},
]


# ═══════════════════════════════════════════════════════════════════════════
# CHART GENERATION HELPERS (Matplotlib -> PNG for PDF embedding)
# ═══════════════════════════════════════════════════════════════════════════

def _create_chart_image(fig: plt.Figure, dpi: int = 150) -> Image:
    """Convert a matplotlib figure to a ReportLab Image for PDF embedding."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor="white", edgecolor="none", transparent=False)
    buf.seek(0)
    img = Image(buf, width=6.5 * inch, height=3.2 * inch)
    plt.close(fig)
    return img


def _chart_recommendation_distribution(df: pd.DataFrame) -> Image:
    """Pie chart: pricing recommendation distribution."""
    rec_col = "recommendation"
    if rec_col not in df.columns:
        return None

    fig, ax = plt.subplots(figsize=(6.5, 3.2), facecolor="white")
    counts = df[rec_col].value_counts()
    labels = counts.index.tolist()
    sizes = counts.values.tolist()
    colors_pie = {"Increase": "#2e7d32", "Decrease": "#c62828", "Maintain": "#1565c0"}
    pie_colors = [colors_pie.get(l, "#757575") for l in labels]
    explode = [0.05 if s == max(sizes) else 0 for s in sizes]

    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, autopct="%1.1f%%", startangle=90,
        colors=pie_colors, explode=explode, shadow=False,
        wedgeprops={"linewidth": 1, "edgecolor": "white"},
        textprops={"fontsize": 9},
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_fontweight("bold")
        at.set_color("white")

    # Legend
    legend_labels = [f"{l} ({s} products)" for l, s in zip(labels, sizes)]
    ax.legend(wedges, legend_labels, loc="center left",
              bbox_to_anchor=(1, 0.5), fontsize=8, frameon=False)

    ax.set_title("Pricing Recommendation Distribution", fontsize=13,
                 fontweight="bold", color="#1a237e", pad=12, loc="center")
    return _create_chart_image(fig)


def _chart_revenue_by_category(df: pd.DataFrame) -> Image:
    """Horizontal bar chart: revenue by category."""
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    cat_col = "category"
    if rev_col not in df.columns or cat_col not in df.columns:
        return None

    cat_rev = df.groupby(cat_col)[rev_col].sum().sort_values(ascending=True)

    fig, ax = plt.subplots(figsize=(6.5, 3.2), facecolor="white")
    y_pos = range(len(cat_rev))
    bars = ax.barh(y_pos, cat_rev.values, height=0.6,
                   color=CHART_COLORS_SEQ[:len(cat_rev)][::-1],
                   edgecolor="white", linewidth=0.5)

    # Add value labels
    for i, (bar, val) in enumerate(zip(bars, cat_rev.values)):
        ax.text(val + max(cat_rev.values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"${val:,.0f}", va="center", fontsize=8, color="#333")

    ax.set_yticks(y_pos)
    ax.set_yticklabels(cat_rev.index, fontsize=9)
    ax.set_xlabel("Revenue ($)", fontsize=9, color="#666")
    ax.set_title("Revenue by Category", fontsize=13, fontweight="bold",
                 color="#1a237e", pad=12, loc="center")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#ddd")
    ax.spines["bottom"].set_color("#ddd")
    ax.tick_params(colors="#666", labelsize=8)
    ax.set_axisbelow(True)
    ax.grid(axis="x", alpha=0.2, color="#999")

    return _create_chart_image(fig)


def _chart_risk_distribution(df: pd.DataFrame) -> Image:
    """Donut/bar chart: risk level distribution."""
    risk_col = "risk_level" if "risk_level" in df.columns else "risk_category"
    if risk_col not in df.columns:
        return None

    order = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    counts = df[risk_col].value_counts()
    data = {l: counts.get(l, 0) for l in order if counts.get(l, 0) > 0}
    labels = list(data.keys())
    sizes = list(data.values())
    colors_risk = {"LOW": "#2e7d32", "MEDIUM": "#f9a825",
                   "HIGH": "#e65100", "CRITICAL": "#c62828"}
    pie_colors = [colors_risk.get(l, "#757575") for l in labels]

    fig, ax = plt.subplots(figsize=(6.5, 3.2), facecolor="white")
    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, autopct="%1.1f%%", startangle=90,
        colors=pie_colors, pctdistance=0.75,
        wedgeprops={"linewidth": 1, "edgecolor": "white"},
        textprops={"fontsize": 9},
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_fontweight("bold")
        at.set_color("white")

    # Draw centre circle for donut effect
    centre_circle = plt.Circle((0, 0), 0.55, fc="white", linewidth=0)
    ax.add_artist(centre_circle)
    ax.text(0, 0, f"{len(df)}", ha="center", va="center",
            fontsize=18, fontweight="bold", color="#1a237e")
    ax.text(0, -0.08, "Products", ha="center", va="center",
            fontsize=7, color="#666")

    legend_labels = [f"{l} ({s})" for l, s in zip(labels, sizes)]
    ax.legend(wedges, legend_labels, loc="center left",
              bbox_to_anchor=(1, 0.5), fontsize=8, frameon=False,
              title="Risk Level", title_fontsize=9)

    ax.set_title("Risk Distribution", fontsize=13, fontweight="bold",
                 color="#1a237e", pad=12, loc="center")
    return _create_chart_image(fig)


def _chart_margin_distribution(df: pd.DataFrame) -> Image:
    """Bar chart: margin distribution by category."""
    margin_col = None
    for c in ["margin_percentage", "profit_margin"]:
        if c in df.columns:
            margin_col = c
            break
    cat_col = "category"
    if margin_col is None or cat_col not in df.columns:
        return None

    cat_margin = df.groupby(cat_col)[margin_col].mean()
    # Convert profit_margin (0-1 scale) to percentage
    if margin_col == "profit_margin":
        cat_margin = cat_margin * 100

    cat_margin = cat_margin.sort_values(ascending=True)

    fig, ax = plt.subplots(figsize=(6.5, 3.2), facecolor="white")
    y_pos = range(len(cat_margin))
    bar_colors = ["#2e7d32" if v >= 30 else "#f9a825" if v >= 15 else "#c62828"
                  for v in cat_margin.values]
    bars = ax.barh(y_pos, cat_margin.values, height=0.6,
                   color=bar_colors, edgecolor="white", linewidth=0.5)

    for i, (bar, val) in enumerate(zip(bars, cat_margin.values)):
        ax.text(val + 0.5, bar.get_y() + bar.get_height() / 2,
                f"{val:.1f}%", va="center", fontsize=8, color="#333",
                fontweight="bold")

    ax.set_yticks(y_pos)
    ax.set_yticklabels(cat_margin.index, fontsize=9)
    ax.set_xlabel("Average Margin (%)", fontsize=9, color="#666")
    ax.set_title("Margin Distribution by Category", fontsize=13,
                 fontweight="bold", color="#1a237e", pad=12, loc="center")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#ddd")
    ax.spines["bottom"].set_color("#ddd")
    ax.tick_params(colors="#666", labelsize=8)
    ax.set_axisbelow(True)
    ax.grid(axis="x", alpha=0.2, color="#999")

    # Add a reference line
    ax.axvline(x=cat_margin.mean(), color="#1a237e", linestyle="--",
               linewidth=1, alpha=0.6, label=f"Avg: {cat_margin.mean():.1f}%")
    ax.legend(fontsize=7, frameon=False, loc="lower right")

    return _create_chart_image(fig)


def _chart_top_revenue_products(df: pd.DataFrame) -> Image:
    """Bar chart: top 10 products by revenue."""
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    name_col = "product_name"
    if rev_col not in df.columns or name_col not in df.columns:
        return None

    top = df.nlargest(10, rev_col)[[name_col, rev_col]].copy()
    top = top.sort_values(rev_col, ascending=True)

    fig, ax = plt.subplots(figsize=(6.5, 3.2), facecolor="white")
    # Truncate long names
    names = [n[:25] + "..." if len(str(n)) > 25 else str(n) for n in top[name_col]]

    colors_bar = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(top))]
    bars = ax.barh(range(len(top)), top[rev_col].values, height=0.6,
                   color=colors_bar[::-1], edgecolor="white", linewidth=0.5)

    for bar, val in zip(bars, top[rev_col].values):
        ax.text(val + max(top[rev_col].values) * 0.01,
                bar.get_y() + bar.get_height() / 2,
                f"${val:,.0f}", va="center", fontsize=7, color="#333")

    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(names[::-1], fontsize=7)
    ax.set_xlabel("Revenue ($)", fontsize=9, color="#666")
    ax.set_title("Top 10 Revenue Products", fontsize=13,
                 fontweight="bold", color="#1a237e", pad=12, loc="center")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#ddd")
    ax.spines["bottom"].set_color("#ddd")
    ax.tick_params(colors="#666", labelsize=7)
    ax.set_axisbelow(True)
    ax.grid(axis="x", alpha=0.2, color="#999")

    return _create_chart_image(fig)


# ═══════════════════════════════════════════════════════════════════════════
# CSV REPORT
# ═══════════════════════════════════════════════════════════════════════════

def generate_csv_report(
    df: pd.DataFrame,
    filename: Optional[str] = None,
    output_dir: Optional[Union[str, Path]] = None,
    index: bool = False,
    **kwargs: Any,
) -> str:
    """
    Generate a CSV report from the analysis DataFrame.

    The CSV includes all analysis columns, making it suitable for
    further processing in Excel, Tableau, or other tools.

    Args:
        df: Analysis DataFrame.
        filename: Output filename (auto-generated if None).
        output_dir: Output directory (default: reports/).
        index: Whether to include the DataFrame index.
        **kwargs: Additional arguments for pd.DataFrame.to_csv().

    Returns:
        Path to the generated CSV file.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if filename is None:
        filename = f"pricing_report_{timestamp}.csv"

    config = AppConfig()
    output_path = (
        Path(output_dir) / filename
        if output_dir
        else config.get_report_path(filename)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Ensure numeric formatting is clean
    export_df = df.copy()
    for col in export_df.select_dtypes(include="float").columns:
        export_df[col] = export_df[col].round(4)

    export_df.to_csv(
        output_path,
        index=index,
        encoding="utf-8-sig",
        **kwargs,
    )

    logger.info(f"CSV report generated: {output_path} ({os.path.getsize(output_path):,} bytes)")
    return str(output_path)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL REPORT  (OpenPyXL)
# ═══════════════════════════════════════════════════════════════════════════

def _excel_header_fill() -> PatternFill:
    return PatternFill(start_color=BRAND_PRIMARY[1:], end_color=BRAND_PRIMARY[1:], fill_type="solid")


def _excel_light_fill() -> PatternFill:
    return PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")


def _excel_green_fill() -> PatternFill:
    return PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")


def _excel_red_fill() -> PatternFill:
    return PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")


def _excel_amber_fill() -> PatternFill:
    return PatternFill(start_color="fff8e1", end_color="fff8e1", fill_type="solid")


def _excel_thin_border() -> Border:
    return Border(
        left=Side(style="thin", color="d0d0d0"),
        right=Side(style="thin", color="d0d0d0"),
        top=Side(style="thin", color="d0d0d0"),
        bottom=Side(style="thin", color="d0d0d0"),
    )


def _write_excel_kpi_sheet(ws: Any, kpi_data: Dict[str, Any]) -> None:
    """
    Write a KPI summary sheet with title, metrics table, and formatting.

    Args:
        ws: OpenPyXL worksheet.
        kpi_data: Dict of KPI metrics (e.g. total_products, total_revenue).
    """
    # Title
    ws["A1"] = "AI Pricing Intelligence Platform — Executive Summary"
    ws["A1"].font = Font(bold=True, size=16, color=BRAND_PRIMARY[1:])
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = Font(size=10, color="666666")

    # KPI Table headers
    headers = ["Metric", "Value", "", ""]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color=BRAND_WHITE[1:], size=11)
        cell.fill = _excel_header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _excel_thin_border()

    ws.merge_cells("A4:B4")
    ws.merge_cells("C4:D4")

    # KPI rows
    kpi_rows = [
        ("Products Analyzed", kpi_data.get("total_products", 0)),
        ("Categories", kpi_data.get("total_categories", 0)),
        ("Total Revenue", f"${kpi_data.get('total_revenue', 0):,.2f}"),
        ("Total Profit", f"${kpi_data.get('total_profit', 0):,.2f}"),
        ("Average Margin", f"{kpi_data.get('avg_margin', 0):.2f}%"),
        ("Average Risk Score", f"{kpi_data.get('avg_risk', 0):.1f}"),
        ("High Risk Products", kpi_data.get("high_risk_count", 0)),
        ("Total Stock Value", f"${kpi_data.get('total_stock_value', 0):,.2f}"),
    ]

    for row_idx, (metric, value) in enumerate(kpi_rows, 5):
        ws.cell(row=row_idx, column=1, value=metric).font = Font(size=11)
        ws.cell(row=row_idx, column=1).border = _excel_thin_border()
        ws.cell(row=row_idx, column=2, value=value).font = Font(size=11, bold=True)
        ws.cell(row=row_idx, column=2).border = _excel_thin_border()
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")

        if row_idx % 2 == 0:
            ws.cell(row=row_idx, column=1).fill = _excel_light_fill()
            ws.cell(row=row_idx, column=2).fill = _excel_light_fill()

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def _write_excel_dataframe_sheet(
    ws: Any,
    df: pd.DataFrame,
    sheet_title: str = "Data",
    freeze_row: int = 2,
) -> None:
    """
    Write a DataFrame to an Excel worksheet with professional formatting.

    Features:
    - Styled header row (dark background, white text)
    - Alternating row colours
    - Auto-adjusted column widths
    - Frozen header row
    - Number formatting for currency columns

    Args:
        ws: OpenPyXL worksheet.
        df: DataFrame to write.
        sheet_title: Worksheet tab name.
        freeze_row: Row to freeze (default 2 = header row).
    """
    # Detect currency / percentage columns for formatting
    currency_keywords = ["price", "revenue", "profit", "cost", "margin", "stock_value"]
    pct_keywords = ["margin", "change", "pct"]

    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True, color=BRAND_WHITE[1:], size=10)
        cell.fill = _excel_header_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _excel_thin_border()

    # Write data
    for row_idx, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=False), 2
    ):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.border = _excel_thin_border()

            # Alternating row colours
            if row_idx % 2 == 0:
                cell.fill = _excel_light_fill()

            # Format detection for column
            col_name_lower = str(df.columns[col_idx - 1]).lower()

            # Currency columns
            if any(kw in col_name_lower for kw in currency_keywords) and not any(
                kw in col_name_lower for kw in pct_keywords
            ):
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

            # Percentage columns
            if any(kw in col_name_lower for kw in pct_keywords):
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")

            # Integer columns
            if isinstance(value, (int, np.integer)) and not isinstance(value, bool):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Freeze header row
    ws.freeze_panes = f"A{freeze_row}"

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = max(
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col_name)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_length + 3, 10), 35
        )


def _add_excel_category_sheet(wb: Any, df: pd.DataFrame) -> None:
    """
    Add a category analysis sheet with grouped metrics and a bar chart.

    Args:
        wb: OpenPyXL workbook.
        df: Analysis DataFrame.
    """
    if "category" not in df.columns:
        return

    ws = wb.create_sheet("Category Analysis")

    # Aggregate by category
    agg_dict: Dict[str, str] = {
        "product_id": "count",
        "current_price": "mean",
        "profit_margin": "mean",
    }

    if "expected_revenue" in df.columns:
        agg_dict["expected_revenue"] = "sum"
    elif "revenue" in df.columns:
        agg_dict["revenue"] = "sum"
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue" if "revenue" in df.columns else None

    if "composite_risk_score" in df.columns:
        agg_dict["composite_risk_score"] = "mean"

    cat_data = df.groupby("category").agg(agg_dict).reset_index()
    col_names = ["Category", "Product Count", "Avg Price", "Avg Margin"]
    if rev_col:
        col_names.append("Total Revenue")
    if "composite_risk_score" in df.columns:
        col_names.append("Avg Risk")
    cat_data.columns = col_names

    cat_data["Avg Margin"] = cat_data["Avg Margin"].apply(lambda x: round(x * 100, 1))
    cat_data["Avg Price"] = cat_data["Avg Price"].apply(lambda x: round(x, 2))

    _write_excel_dataframe_sheet(ws, cat_data, sheet_title="Category Analysis")

    # Add a bar chart for revenue by category
    if rev_col and "Total Revenue" in cat_data.columns:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Total Revenue by Category"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Category"
        chart.style = 10

        data_ref = Reference(ws, min_col=cat_data.columns.get_loc("Total Revenue") + 1,
                             min_row=1, max_row=len(cat_data) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2,
                             max_row=len(cat_data) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4

        # Colour the bars
        series = chart.series[0]
        series.graphicalProperties.solidFill = BRAND_SECONDARY

        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, f"E2")


def _add_excel_risk_sheet(wb: Any, df: pd.DataFrame) -> None:
    """
    Add a risk analysis sheet with risk level breakdown.

    Args:
        wb: OpenPyXL workbook.
        df: Analysis DataFrame.
    """
    risk_level_col = "risk_level" if "risk_level" in df.columns else "risk_category"
    if risk_level_col not in df.columns:
        return

    ws = wb.create_sheet("Risk Analysis")

    # Risk level distribution
    level_counts = df[risk_level_col].value_counts()
    risk_data = pd.DataFrame({
        "Risk Level": level_counts.index,
        "Count": level_counts.values,
    })
    risk_data["Percentage"] = (risk_data["Count"] / risk_data["Count"].sum() * 100).round(1)

    # Order
    risk_data["Risk Level"] = pd.Categorical(
        risk_data["Risk Level"],
        categories=RISK_LEVEL_ORDER,
        ordered=True,
    )
    risk_data = risk_data.sort_values("Risk Level")

    _write_excel_dataframe_sheet(ws, risk_data, sheet_title="Risk Analysis")

    # Pie chart for risk distribution
    chart = PieChart()
    chart.title = "Risk Level Distribution"
    chart.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(risk_data) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(risk_data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Pie colours
    colors_seq = [BRAND_DANGER, "#f44336", BRAND_WARNING, BRAND_ACCENT]
    for i, color in enumerate(colors_seq[:len(risk_data)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart.series[0].data_points.append(pt)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True

    chart.width = 16
    chart.height = 12
    ws.add_chart(chart, f"E2")


def _add_excel_recommendation_sheet(wb: Any, df: pd.DataFrame) -> None:
    """
    Add a pricing recommendations sheet.

    Args:
        wb: OpenPyXL workbook.
        df: Analysis DataFrame.
    """
    if "recommendation" not in df.columns:
        return

    ws = wb.create_sheet("Recommendations")

    rec_counts = df["recommendation"].value_counts().reset_index()
    rec_counts.columns = ["Recommendation", "Count"]

    _write_excel_dataframe_sheet(ws, rec_counts, sheet_title="Recommendations")

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Pricing Recommendations"
    chart.y_axis.title = "Product Count"

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(rec_counts) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(rec_counts) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.style = 10

    colors_map = {
        "Increase": BRAND_ACCENT,
        "Decrease": BRAND_DANGER,
        "Maintain": BRAND_SECONDARY,
    }
    for i in range(len(rec_counts)):
        pt = DataPoint(idx=i)
        rec = rec_counts.iloc[i]["Recommendation"]
        pt.graphicalProperties.solidFill = colors_map.get(rec, "#757575")
        chart.series[0].data_points.append(pt)

    chart.width = 16
    chart.height = 12
    ws.add_chart(chart, f"E2")


def generate_excel_report(
    df: pd.DataFrame,
    insights: Optional[Dict[str, Any]] = None,
    filename: Optional[str] = None,
    include_kpi: bool = True,
    include_category: bool = True,
    include_risk: bool = True,
    include_recommendations: bool = True,
    include_full_data: bool = True,
) -> str:
    """
    Generate a professional multi-sheet Excel report.

    Sheets:
    1. Executive Summary — KPIs and high-level metrics
    2. Product Data — full analysed dataset
    3. Category Analysis — grouped metrics + revenue bar chart
    4. Risk Analysis — risk level distribution + pie chart
    5. Pricing Recommendations — recommendation breakdown + bar chart

    Args:
        df: Analysis DataFrame with all computed columns.
        insights: Dict of KPI metrics (auto-computed if None).
        filename: Output filename (auto-generated if None).
        include_kpi: Include the KPI summary sheet.
        include_category: Include category analysis sheet.
        include_risk: Include risk analysis sheet.
        include_recommendations: Include recommendations sheet.
        include_full_data: Include full product data sheet.

    Returns:
        Path to the generated Excel file.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if filename is None:
        filename = f"pricing_report_{timestamp}.xlsx"

    config = AppConfig()
    filepath = config.get_report_path(filename)

    # Auto-compute insights if not provided
    if insights is None:
        insights = _compute_insights(df)

    wb = openpyxl.Workbook()

    # Sheet 1: Executive Summary
    if include_kpi:
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        _write_excel_kpi_sheet(ws_summary, insights)

    # Sheet 2: Full Product Data
    if include_full_data:
        ws_name = "Product Data" if include_kpi else wb.active
        ws_data = wb.create_sheet("Product Data") if include_kpi else wb.active
        if not include_kpi:
            ws_data.title = "Product Data"
        _write_excel_dataframe_sheet(ws_data, df, sheet_title="Product Data")

    # Sheet 3: Category Analysis
    if include_category:
        _add_excel_category_sheet(wb, df)

    # Sheet 4: Risk Analysis
    if include_risk:
        _add_excel_risk_sheet(wb, df)

    # Sheet 5: Recommendations
    if include_recommendations:
        _add_excel_recommendation_sheet(wb, df)

    wb.save(str(filepath))
    logger.info(
        f"Excel report generated: {filepath} "
        f"({os.path.getsize(filepath):,} bytes, {len(wb.sheetnames)} sheets: {wb.sheetnames})"
    )
    return str(filepath)


# ═══════════════════════════════════════════════════════════════════════════
# PDF REPORT — IMPROVED EXECUTIVE-LEVEL BUSINESS REPORT
# ═══════════════════════════════════════════════════════════════════════════

def _get_pdf_styles() -> Dict[str, ParagraphStyle]:
    """
    Create and return a dictionary of custom paragraph styles for the PDF report.

    Returns:
        Dict of style name -> ParagraphStyle.
    """
    styles = getSampleStyleSheet()

    custom_styles = {
        "ReportTitle": ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontSize=28,
            leading=34,
            spaceAfter=4,
            textColor=colors.HexColor(BRAND_PRIMARY),
            alignment=TA_CENTER,
        ),
        "ReportSubtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontSize=13,
            leading=17,
            spaceAfter=6,
            textColor=colors.HexColor("#455a64"),
            alignment=TA_CENTER,
        ),
        "CoverSubtitle": ParagraphStyle(
            "CoverSubtitle",
            parent=styles["Normal"],
            fontSize=11,
            leading=14,
            spaceAfter=4,
            textColor=colors.HexColor("#78909c"),
            alignment=TA_CENTER,
        ),
        "SectionHeader": ParagraphStyle(
            "SectionHeader",
            parent=styles["Heading2"],
            fontSize=16,
            leading=20,
            spaceBefore=18,
            spaceAfter=10,
            textColor=colors.HexColor(BRAND_PRIMARY),
            borderWidth=0,
            borderPadding=0,
        ),
        "SubSectionHeader": ParagraphStyle(
            "SubSectionHeader",
            parent=styles["Heading3"],
            fontSize=13,
            leading=17,
            spaceBefore=12,
            spaceAfter=6,
            textColor=colors.HexColor(BRAND_SECONDARY),
        ),
        "KPIValue": ParagraphStyle(
            "KPIValue",
            parent=styles["Normal"],
            fontSize=24,
            leading=28,
            textColor=colors.HexColor(BRAND_ACCENT),
            alignment=TA_CENTER,
        ),
        "KPIValueRed": ParagraphStyle(
            "KPIValueRed",
            parent=styles["Normal"],
            fontSize=24,
            leading=28,
            textColor=colors.HexColor(BRAND_DANGER),
            alignment=TA_CENTER,
        ),
        "KPILabel": ParagraphStyle(
            "KPILabel",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#616161"),
            alignment=TA_CENTER,
        ),
        "KPIExplanation": ParagraphStyle(
            "KPIExplanation",
            parent=styles["Normal"],
            fontSize=7,
            leading=9,
            textColor=colors.HexColor("#78909c"),
            alignment=TA_CENTER,
        ),
        "BusinessInsight": ParagraphStyle(
            "BusinessInsight",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=6,
            leftIndent=10,
            textColor=colors.HexColor("#37474f"),
        ),
        "BodyText2": ParagraphStyle(
            "BodyText2",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=6,
        ),
        "BodyTextSmall": ParagraphStyle(
            "BodyTextSmall",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            spaceAfter=4,
            textColor=colors.HexColor("#546e7a"),
        ),
        "BulletText": ParagraphStyle(
            "BulletText",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=4,
            leftIndent=15,
            bulletIndent=5,
        ),
        "FooterStyle": ParagraphStyle(
            "FooterStyle",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor("#9e9e9e"),
            alignment=TA_CENTER,
        ),
        "Disclaimer": ParagraphStyle(
            "Disclaimer",
            parent=styles["Normal"],
            fontSize=7,
            leading=10,
            textColor=colors.HexColor("#bdbdbd"),
            alignment=TA_CENTER,
        ),
        "RecommendationReason": ParagraphStyle(
            "RecommendationReason",
            parent=styles["Normal"],
            fontSize=8,
            leading=11,
            spaceAfter=4,
            textColor=colors.HexColor("#546e7a"),
        ),
        "CategoryCommentary": ParagraphStyle(
            "CategoryCommentary",
            parent=styles["Normal"],
            fontSize=9,
            leading=13,
            spaceAfter=6,
            textColor=colors.HexColor("#37474f"),
            leftIndent=5,
        ),
        "ActionHeader": ParagraphStyle(
            "ActionHeader",
            parent=styles["Normal"],
            fontSize=11,
            leading=14,
            spaceBefore=6,
            spaceAfter=4,
            textColor=colors.HexColor(BRAND_ACCENT),
        ),
    }

    # Register with the stylesheet so Paragraph can find them
    for style in custom_styles.values():
        styles.add(style)

    return styles


def _compute_insights(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Compute a standard set of KPI insights from a DataFrame.

    Args:
        df: Analysis DataFrame.

    Returns:
        Dict of KPI metrics.
    """
    insights: Dict[str, Any] = {}

    # Basic counts
    insights["total_products"] = len(df)
    insights["total_categories"] = int(df["category"].nunique()) if "category" in df.columns else 0

    # Revenue
    for col in ["expected_revenue", "revenue"]:
        if col in df.columns:
            insights["total_revenue"] = round(float(df[col].sum()), 2)
            break
    if "total_revenue" not in insights:
        if "current_price" in df.columns and "sales_volume" in df.columns:
            insights["total_revenue"] = round(float((df["current_price"] * df["sales_volume"]).sum()), 2)
        else:
            insights["total_revenue"] = 0

    # Current revenue (for delta calculations)
    if "current_price" in df.columns and "sales_volume" in df.columns:
        insights["current_revenue"] = round(float((df["current_price"] * df["sales_volume"]).sum()), 2)
    else:
        insights["current_revenue"] = insights["total_revenue"]

    # Profit
    for col in ["expected_profit"]:
        if col in df.columns:
            insights["total_profit"] = round(float(df[col].sum()), 2)
            break
    if "total_profit" not in insights:
        insights["total_profit"] = 0

    # Margin
    for col in ["margin_percentage", "profit_margin"]:
        if col in df.columns:
            if col == "profit_margin":
                insights["avg_margin"] = round(float(df[col].mean() * 100), 2)
            else:
                insights["avg_margin"] = round(float(df[col].mean()), 2)
            break
    if "avg_margin" not in insights:
        insights["avg_margin"] = 0

    # Risk
    risk_col = "composite_risk_score"
    if risk_col in df.columns:
        insights["avg_risk"] = round(float(df[risk_col].mean()), 1)
        insights["high_risk_count"] = int((df[risk_col] >= 70).sum())
        insights["moderate_risk_count"] = int(((df[risk_col] >= 40) & (df[risk_col] < 70)).sum())
    else:
        insights["avg_risk"] = 0
        insights["high_risk_count"] = 0
        insights["moderate_risk_count"] = 0

    # Inventory
    if "stock_value" in df.columns:
        insights["total_stock_value"] = round(float(df["stock_value"].sum()), 2)

    # Recommendations
    if "recommendation" in df.columns:
        rec_counts = df["recommendation"].value_counts()
        for rec in ["Increase", "Decrease", "Maintain"]:
            insights[f"rec_{rec.lower()}"] = int(rec_counts.get(rec, 0))
        # Compute lifts
        if "expected_revenue" in df.columns:
            current_rev = insights.get("current_revenue", 0)
            insights["revenue_lift"] = round(insights["total_revenue"] - current_rev, 2)
        if "expected_profit" in df.columns and "current_price" in df.columns and "cost_price" in df.columns:
            current_profit = float(((df["current_price"] - df["cost_price"]) * df["sales_volume"]).sum())
            insights["profit_lift"] = round(insights["total_profit"] - current_profit, 2)

    return insights


def _insights_from_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    """Alias for _compute_insights for backwards compatibility."""
    return _compute_insights(df)


def _generate_narrative_insights(df: pd.DataFrame, insights: Dict[str, Any]) -> List[str]:
    """
    Generate narrative business insights from the data.

    Args:
        df: Analysis DataFrame.
        insights: KPI insights dict.

    Returns:
        List of narrative insight strings.
    """
    narratives = []

    # Portfolio overview
    total = insights.get("total_products", len(df))
    categories = insights.get("total_categories", 0)
    narratives.append(
        f"This report analyzes <b>{total}</b> products across <b>{categories}</b> "
        f"categories to identify pricing opportunities, revenue potential, "
        f"profitability trends, and business risks."
    )

    # Pricing insights
    if "recommendation" in df.columns:
        rec_counts = df["recommendation"].value_counts()
        inc_pct = rec_counts.get("Increase", 0) / total * 100 if total > 0 else 0
        dec_pct = rec_counts.get("Decrease", 0) / total * 100 if total > 0 else 0
        maint_pct = rec_counts.get("Maintain", 0) / total * 100 if total > 0 else 0

        if dec_pct > 30:
            narratives.append(
                f"<b>Pricing Insight:</b> {dec_pct:.1f}% of products require price reductions, "
                f"suggesting possible overpricing or competitive pressure in the portfolio."
            )
        if inc_pct > 20:
            narratives.append(
                f"<b>Pricing Opportunity:</b> {inc_pct:.1f}% of products have opportunities for "
                f"price increases, representing potential revenue upside."
            )
        if maint_pct > 40:
            narratives.append(
                f"<b>Price Stability:</b> {maint_pct:.1f}% of products are appropriately priced "
                f"and should maintain current levels."
            )

    # Revenue concentration
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    if rev_col in df.columns:
        top10 = df.nlargest(10, rev_col)[rev_col].sum()
        total_rev = insights.get("total_revenue", df[rev_col].sum())
        if total_rev > 0:
            conc_pct = top10 / total_rev * 100
            if conc_pct > 50:
                narratives.append(
                    f"<b>Revenue Concentration:</b> Top 10 products contribute "
                    f"{conc_pct:.1f}% of total revenue, indicating significant "
                    f"revenue concentration risk."
                )
            else:
                narratives.append(
                    f"<b>Revenue Distribution:</b> Revenue is reasonably distributed "
                    f"across the portfolio, with top 10 products contributing "
                    f"{conc_pct:.1f}% of total revenue."
                )

    # Margin insights
    margin_col = None
    for c in ["margin_percentage", "profit_margin"]:
        if c in df.columns:
            margin_col = c
            break
    if margin_col:
        avg_margin = insights.get("avg_margin", 0)
        if margin_col == "profit_margin":
            avg_margin = avg_margin  # already converted in insights
        if avg_margin < 20:
            narratives.append(
                f"<b>Margin Pressure:</b> Average margin of {avg_margin:.1f}% is below "
                f"the healthy threshold, suggesting margin improvement opportunities."
            )
        elif avg_margin > 40:
            narratives.append(
                f"<b>Strong Margins:</b> Average margin of {avg_margin:.1f}% indicates "
                f"healthy portfolio profitability."
            )
        else:
            narratives.append(
                f"<b>Margin Profile:</b> Average portfolio margin of {avg_margin:.1f}% "
                f"is within acceptable range."
            )

    # Risk insights
    if "composite_risk_score" in df.columns:
        avg_risk = insights.get("avg_risk", 0)
        high_risk = insights.get("high_risk_count", 0)
        if high_risk > 3:
            narratives.append(
                f"<b>Risk Watch:</b> {high_risk} products require immediate business attention "
                f"with elevated risk scores (avg: {avg_risk:.1f})."
            )
        elif high_risk > 0:
            narratives.append(
                f"<b>Risk Profile:</b> {high_risk} product(s) identified as high-risk "
                f"warranting periodic review."
            )

    # Category insights
    if "category" in df.columns:
        cat_count = df["category"].nunique()
        if cat_count > 1:
            # Find top category by revenue
            if rev_col in df.columns:
                top_cat = df.groupby("category")[rev_col].sum().idxmax()
                top_cat_rev = df.groupby("category")[rev_col].sum().max()
                total_rev = insights.get("total_revenue", 0)
                cat_share = top_cat_rev / total_rev * 100 if total_rev > 0 else 0
                narratives.append(
                    f"<b>Category Performance:</b> '{top_cat}' leads with "
                    f"{cat_share:.1f}% of portfolio revenue."
                )

                # Category with highest risk
                if "composite_risk_score" in df.columns:
                    high_risk_cat = df.groupby("category")["composite_risk_score"].mean().idxmax()
                    high_risk_val = df.groupby("category")["composite_risk_score"].mean().max()
                    narratives.append(
                        f"<b>Category Risk:</b> '{high_risk_cat}' exhibits the highest "
                        f"average risk score ({high_risk_val:.1f}) in the portfolio."
                    )

    # Inventory/demand observations
    if "demand_trend" in df.columns:
        declining = (df["demand_trend"] < 0.5).sum()
        if declining > 0:
            pct_declining = declining / total * 100
            narratives.append(
                f"<b>Demand Signal:</b> {pct_declining:.1f}% of products show declining "
                f"demand trends, requiring inventory planning attention."
            )

    return narratives


def _build_pdf_cover_page(story: List[Any], styles: Dict[str, ParagraphStyle],
                           insights: Dict[str, Any]) -> None:
    """
    Build a professional cover page for the PDF report.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        insights: KPI insights dict.
    """
    # Top brand bar
    story.append(Spacer(1, 0.1 * inch))
    story.append(HRFlowable(
        width="100%", thickness=4, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=40, spaceBefore=0,
    ))

    # Title area
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("AI Pricing Intelligence", styles["ReportTitle"]))
    story.append(Paragraph("Executive Pricing Analysis Report", styles["ReportTitle"]))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        "Automated Pricing Optimization & Portfolio Insights",
        styles["ReportSubtitle"],
    ))
    story.append(Spacer(1, 0.25 * inch))

    # Divider
    story.append(HRFlowable(
        width="60%", thickness=1.5, color=colors.HexColor(BRAND_SECONDARY),
        spaceAfter=20, spaceBefore=10,
    ))

    # Report metadata
    timestamp = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    meta_data = [
        f"Report Generated: {timestamp}",
        f"Products Analyzed: {insights.get('total_products', 'N/A')}",
        f"Categories Covered: {insights.get('total_categories', 'N/A')}",
    ]
    for line in meta_data:
        story.append(Paragraph(line, styles["CoverSubtitle"]))

    story.append(Spacer(1, 0.4 * inch))

    # Quick KPI snapshot box
    story.append(Paragraph("QUICK PERFORMANCE SNAPSHOT", styles["KPILabel"]))
    story.append(Spacer(1, 0.1 * inch))

    # KPI row on cover
    cover_kpis = [
        ("Total Revenue", f"${insights.get('total_revenue', 0):,.0f}"),
        ("Total Profit", f"${insights.get('total_profit', 0):,.0f}"),
        ("Avg Margin", f"{insights.get('avg_margin', 0):.1f}%"),
    ]
    if insights.get("high_risk_count", 0) > 0:
        cover_kpis.append(("High Risk", str(insights.get("high_risk_count", 0))))

    kpi_cells = []
    for label, value in cover_kpis:
        cell_text = (
            f"<b><font size='22' color='#1a237e'>{value}</font></b><br/>"
            f"<font size='8' color='#78909c'>{label}</font>"
        )
        kpi_cells.append(Paragraph(cell_text, styles["Normal"]))

    kpi_table = Table(
        [kpi_cells],
        colWidths=[1.6 * inch] * len(kpi_cells),
    )
    kpi_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
        ("BOX", (0, 0), (-1, 0), 1, colors.HexColor("#e0e0e0")),
    ]))
    story.append(kpi_table)

    story.append(Spacer(1, 0.5 * inch))

    # Bottom bar
    story.append(HRFlowable(
        width="100%", thickness=2, color=colors.HexColor(BRAND_SECONDARY),
        spaceAfter=20, spaceBefore=20,
    ))
    story.append(Paragraph(
        "CONFIDENTIAL — Prepared for Internal Business Use",
        styles["Disclaimer"],
    ))
    story.append(Paragraph(
        "AI Pricing Intelligence Platform | Automated Reporting Module",
        styles["Disclaimer"],
    ))


def _build_pdf_executive_summary(story: List[Any], styles: Dict[str, ParagraphStyle],
                                  insights: Dict[str, Any], narratives: List[str]) -> None:
    """
    Build the executive summary section with narrative insights.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        insights: KPI insights dict.
        narratives: List of narrative insight strings.
    """
    # Section header with underline bar
    story.append(Paragraph("1. Executive Summary", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    # Main narrative paragraph
    if narratives:
        story.append(Paragraph(narratives[0], styles["BusinessInsight"]))
        story.append(Spacer(1, 0.05 * inch))

    # Portfolio health summary
    health_items = []
    if "total_revenue" in insights:
        health_items.append(
            f"<b>Portfolio Value:</b> Total revenue of <b>${insights['total_revenue']:,.0f}</b> "
            f"across {insights.get('total_products', 0)} products."
        )
    if "total_profit" in insights:
        health_items.append(
            f"<b>Profitability:</b> Total profit estimated at <b>${insights['total_profit']:,.0f}</b> "
            f"with average margin of <b>{insights.get('avg_margin', 0):.1f}%</b>."
        )
    if "composite_risk_score" in df_global_placeholder if False else True:
        pass  # We'll add risk summary from insights
    if "high_risk_count" in insights and insights["high_risk_count"] > 0:
        health_items.append(
            f"<b>Risk Profile:</b> <b>{insights['high_risk_count']}</b> product(s) flagged as "
            f"high-risk, requiring management attention."
        )

    # We need access to df for these narrative insights
    for item in health_items:
        story.append(Paragraph(f"• {item}", styles["BulletText"]))

    story.append(Spacer(1, 0.1 * inch))

    # Additional narrative insights (skip the first one which was the intro)
    for narrative in narratives[1:]:
        story.append(Paragraph(narrative, styles["BusinessInsight"]))


def _build_pdf_kpi_dashboard(story: List[Any], styles: Dict[str, ParagraphStyle],
                               insights: Dict[str, Any]) -> None:
    """
    Build the Executive KPI Dashboard with explanation cards.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        insights: KPI insights dict.
    """
    story.append(Paragraph("2. Executive KPI Dashboard", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    # Define KPIs with explanations
    kpi_definitions = [
        {
            "value": f"${insights.get('revenue_lift', 0):,.0f}",
            "label": "Revenue Lift",
            "explanation": "Expected increase in total revenue after applying recommended pricing changes.",
            "color": BRAND_ACCENT if insights.get("revenue_lift", 0) >= 0 else BRAND_DANGER,
        },
        {
            "value": f"${insights.get('profit_lift', 0):,.0f}",
            "label": "Profit Lift",
            "explanation": "Estimated profit improvement from optimized pricing decisions.",
            "color": BRAND_ACCENT if insights.get("profit_lift", 0) >= 0 else BRAND_DANGER,
        },
        {
            "value": f"{insights.get('avg_margin', 0):.1f}%",
            "label": "Average Margin",
            "explanation": "Average profitability across the analyzed portfolio.",
            "color": BRAND_PRIMARY,
        },
        {
            "value": f"{insights.get('total_revenue', 0):,.0f}",
            "label": "Total Revenue",
            "explanation": f"Aggregate revenue from {insights.get('total_products', 0)} products across all categories.",
            "color": BRAND_PRIMARY,
        },
        {
            "value": f"${insights.get('total_profit', 0):,.0f}",
            "label": "Total Profit",
            "explanation": "Combined profitability after accounting for costs and pricing optimization.",
            "color": BRAND_SECONDARY,
        },
        {
            "value": str(insights.get("high_risk_count", 0)) + " products",
            "label": "High Risk Products",
            "explanation": "Products requiring immediate business attention due to elevated risk scores.",
            "color": BRAND_DANGER if insights.get("high_risk_count", 0) > 0 else BRAND_ACCENT,
        },
    ]

    # Render in 2 rows of 3
    for row_start in range(0, 6, 3):
        row_kpis = kpi_definitions[row_start:row_start + 3]
        row_cells = []
        for kpi in row_kpis:
            val_style_name = "KPIValue"
            cell_text = (
                f"<b><font size='22' color='{kpi['color']}'>{kpi['value']}</font></b><br/>"
                f"<font size='9' color='#37474f'><b>{kpi['label']}</b></font><br/>"
                f"<font size='7' color='#78909c'>{kpi['explanation']}</font>"
            )
            row_cells.append(Paragraph(cell_text, styles["Normal"]))

        kpi_table = Table(
            [row_cells],
            colWidths=[2.1 * inch] * len(row_cells),
        )
        kpi_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, 0), 0.8, colors.HexColor("#e0e0e0")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fafafa")),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.08 * inch))
    story.append(Spacer(1, 0.1 * inch))


def _build_pdf_business_insights(story: List[Any], styles: Dict[str, ParagraphStyle],
                                  df: pd.DataFrame, insights: Dict[str, Any],
                                  narratives: List[str]) -> None:
    """
    Build detailed business insights section with intelligent commentary.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
        insights: KPI insights dict.
        narratives: Narrative insights (skip first which is intro).
    """
    story.append(Paragraph("3. Business Insights", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    # Pricing insights
    story.append(Paragraph("Pricing Insights", styles["SubSectionHeader"]))
    if "recommendation" in df.columns:
        rec_counts = df["recommendation"].value_counts()
        total = len(df)
        inc_pct = rec_counts.get("Increase", 0) / total * 100 if total > 0 else 0
        dec_pct = rec_counts.get("Decrease", 0) / total * 100 if total > 0 else 0
        maint_pct = rec_counts.get("Maintain", 0) / total * 100 if total > 0 else 0

        insights_text = [
            f"• <b>{inc_pct:.1f}%</b> of products ({rec_counts.get('Increase', 0)}) identified for "
            f"<b>price increases</b> — representing direct revenue growth opportunities.",
            f"• <b>{dec_pct:.1f}%</b> of products ({rec_counts.get('Decrease', 0)}) require "
            f"<b>price reductions</b> — suggesting competitive pressure or overpricing.",
            f"• <b>{maint_pct:.1f}%</b> of products ({rec_counts.get('Maintain', 0)}) should "
            f"<b>maintain current prices</b> — currently well-positioned in market.",
        ]
        for t in insights_text:
            story.append(Paragraph(t, styles["BulletText"]))

    story.append(Spacer(1, 0.08 * inch))

    # Revenue insights
    story.append(Paragraph("Revenue Insights", styles["SubSectionHeader"]))
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    if rev_col in df.columns:
        top5 = df.nlargest(5, rev_col)
        top5_rev = top5[rev_col].sum()
        total_rev = insights.get("total_revenue", 0)
        top5_pct = top5_rev / total_rev * 100 if total_rev > 0 else 0
        story.append(Paragraph(
            f"• Top 5 products contribute <b>{top5_pct:.1f}%</b> of total revenue "
            f"(${top5_rev:,.0f} of ${total_rev:,.0f}).",
            styles["BulletText"],
        ))
        if top5_pct > 40:
            story.append(Paragraph(
                "• <b>Revenue concentration risk detected.</b> Diversification strategies "
                "should be considered to reduce dependency on top performers.",
                styles["BulletText"],
            ))

    story.append(Spacer(1, 0.08 * inch))

    # Category insights
    if "category" in df.columns and rev_col in df.columns:
        story.append(Paragraph("Category Insights", styles["SubSectionHeader"]))
        cat_stats = df.groupby("category").agg(
            product_count=("product_id", "count"),
            total_revenue=(rev_col, "sum"),
            avg_risk=("composite_risk_score", "mean") if "composite_risk_score" in df.columns else ("product_id", "count"),
        ).reset_index()
        if "composite_risk_score" not in df.columns:
            cat_stats = cat_stats.drop(columns=["avg_risk"], errors="ignore")

        total_rev = insights.get("total_revenue", 0)
        for _, row in cat_stats.iterrows():
            share = row["total_revenue"] / total_rev * 100 if total_rev > 0 else 0
            risk_note = ""
            if "avg_risk" in row and pd.notna(row.get("avg_risk")):
                risk_level = "elevated" if row["avg_risk"] > 50 else "moderate" if row["avg_risk"] > 30 else "low"
                risk_note = f" with <b>{risk_level}</b> risk ({row['avg_risk']:.0f})"
            story.append(Paragraph(
                f"• <b>{row['category']}</b>: {int(row['product_count'])} products, "
                f"{share:.1f}% of revenue (${row['total_revenue']:,.0f}){risk_note}.",
                styles["BulletText"],
            ))

    story.append(Spacer(1, 0.1 * inch))


def _build_pdf_pricing_action_plan(story: List[Any], styles: Dict[str, ParagraphStyle],
                                    df: pd.DataFrame) -> None:
    """
    Build the Pricing Action Plan section with 3 groups.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
    """
    if "recommendation" not in df.columns:
        return

    story.append(Paragraph("4. Pricing Action Plan", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    # Helper: build action table for a recommendation type
    def _add_action_group(rec_type: str, title: str, color_hex: str, icon: str):
        group_df = df[df["recommendation"] == rec_type].copy()
        if group_df.empty:
            return

        change_col = "price_change_pct" if "price_change_pct" in df.columns else None

        # Sort by impact
        if change_col and rec_type == "Increase":
            group_df = group_df.nlargest(min(8, len(group_df)), change_col)
        elif change_col and rec_type == "Decrease":
            group_df = group_df.nsmallest(min(8, len(group_df)), change_col)
        else:
            group_df = group_df.head(8)

        story.append(Paragraph(f"{icon} {title}", styles["SubSectionHeader"]))

        # Build table
        headers = ["Product", "Current", "Recommended", "Change", "Impact"]
        table_data = [headers]

        rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"

        for _, row in group_df.iterrows():
            prod_name = str(row.get("product_name", "Unknown"))[:25]
            curr = row.get("current_price", 0)
            optimal = row.get("optimal_price", 0)
            change = row.get("price_change_pct", 0)
            impact = row.get(rev_col, 0) if rev_col else 0

            change_str = f"+{change:.1f}%" if change > 0 else f"{change:.1f}%"
            table_data.append([
                prod_name,
                f"${curr:.2f}",
                f"${optimal:.2f}",
                change_str,
                f"${impact:,.0f}" if impact else "—",
            ])

        if len(table_data) > 1:
            # Add reasoning row
            reason_row = None
            if rec_type == "Increase":
                reason_row = [
                    "", "", "", "", "Strong margin opportunity with acceptable competitive positioning."
                ]
            elif rec_type == "Decrease":
                reason_row = [
                    "", "", "", "", "Competitive pressure / demand sensitivity detected."
                ]
            else:
                reason_row = [
                    "", "", "", "", "Current pricing is near-optimal for market conditions."
                ]
            if reason_row:
                table_data.append(reason_row)

            action_table = Table(
                table_data,
                colWidths=[1.6 * inch, 0.8 * inch, 0.9 * inch, 0.7 * inch, 1.2 * inch],
            )
            action_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(color_hex)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -2), 8),
                ("FONTSIZE", (0, -1), (-1, -1), 7),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Oblique"),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#78909c")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -2), 0.4, colors.HexColor("#e0e0e0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2),
                 [colors.white, colors.HexColor("#f5f5f5")]),
                ("LINEBELOW", (0, -2), (-1, -2), 0.8, colors.HexColor(color_hex)),
                ("SPAN", (0, -1), (-4, -1)),
            ]))
            story.append(action_table)
            story.append(Spacer(1, 0.12 * inch))
        else:
            story.append(Paragraph(
                f"No products in this category.",
                styles["BodyTextSmall"],
            ))

    # Group 1: Increase
    _add_action_group("Increase",
                      "Immediate Price Increase Opportunities",
                      BRAND_ACCENT, "▲")

    # Group 2: Decrease
    _add_action_group("Decrease",
                      "Immediate Price Reduction Required",
                      BRAND_DANGER, "▼")

    # Group 3: Maintain
    _add_action_group("Maintain",
                      "Maintain Current Pricing",
                      BRAND_SECONDARY, "●")


def _build_pdf_visualizations(story: List[Any], styles: Dict[str, ParagraphStyle],
                               df: pd.DataFrame) -> None:
    """
    Add professional charts to the report.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
    """
    story.append(Paragraph("5. Visual Analytics", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    # Row 1: Recommendation distribution + Revenue by category
    rec_chart = _chart_recommendation_distribution(df)
    if rec_chart:
        story.append(Paragraph("Pricing Recommendation Distribution", styles["SubSectionHeader"]))
        story.append(rec_chart)
        story.append(Spacer(1, 0.1 * inch))

    rev_chart = _chart_revenue_by_category(df)
    if rev_chart:
        story.append(Paragraph("Revenue by Category", styles["SubSectionHeader"]))
        story.append(rev_chart)
        story.append(Spacer(1, 0.1 * inch))

    risk_chart = _chart_risk_distribution(df)
    if risk_chart:
        story.append(Paragraph("Risk Distribution", styles["SubSectionHeader"]))
        story.append(risk_chart)
        story.append(Spacer(1, 0.1 * inch))

    margin_chart = _chart_margin_distribution(df)
    if margin_chart:
        story.append(Paragraph("Margin Distribution by Category", styles["SubSectionHeader"]))
        story.append(margin_chart)
        story.append(Spacer(1, 0.1 * inch))

    top_rev_chart = _chart_top_revenue_products(df)
    if top_rev_chart:
        story.append(Paragraph("Top Revenue Products", styles["SubSectionHeader"]))
        story.append(top_rev_chart)


def _build_pdf_category_performance(story: List[Any], styles: Dict[str, ParagraphStyle],
                                     df: pd.DataFrame, insights: Dict[str, Any]) -> None:
    """
    Build a business-focused category performance section with commentary.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
        insights: KPI insights dict.
    """
    if "category" not in df.columns:
        return

    story.append(Paragraph("6. Category Performance", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_PRIMARY),
        spaceAfter=10, spaceBefore=0,
    ))

    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    margin_col = None
    for c in ["margin_percentage", "profit_margin"]:
        if c in df.columns:
            margin_col = c
            break
    risk_col = "composite_risk_score" if "composite_risk_score" in df.columns else None

    # Aggregate category data
    agg_dict = {"product_id": "count"}
    if rev_col:
        agg_dict[rev_col] = "sum"
    if margin_col:
        agg_dict[margin_col] = "mean"
    if risk_col:
        agg_dict[risk_col] = "mean"

    cat_data = df.groupby("category").agg(agg_dict).reset_index()
    total_rev = insights.get("total_revenue", 0)

    # Build display table
    headers = ["Category", "Products", "Revenue", "Rev Share", "Avg Margin", "Avg Risk"]
    if not risk_col:
        headers = headers[:-1]

    table_data = [headers]

    for _, row in cat_data.iterrows():
        rev_share = row[rev_col] / total_rev * 100 if total_rev > 0 else 0
        margin_val = row.get(margin_col, 0)
        if margin_col == "profit_margin":
            margin_val = margin_val * 100

        row_data = [
            str(row["category"]),
            str(int(row["product_id"])),
            f"${row[rev_col]:,.0f}",
            f"{rev_share:.1f}%",
            f"{margin_val:.1f}%",
        ]
        if risk_col:
            row_data.append(f"{row.get(risk_col, 0):.1f}")
        table_data.append(row_data)

    if len(table_data) > 1:
        col_widths = [1.2 * inch, 0.6 * inch, 1.0 * inch, 0.7 * inch, 0.8 * inch]
        if risk_col:
            col_widths.append(0.7 * inch)
        # Adjust for remaining space
        remaining = 6.5 - sum(col_widths)
        col_widths[0] += remaining

        cat_table = Table(table_data, colWidths=col_widths)
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_PRIMARY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e0e0e0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.1 * inch))

    # Generate commentary for each category
    story.append(Paragraph("Category Commentary", styles["SubSectionHeader"]))
    for _, row in cat_data.iterrows():
        margin_val = row.get(margin_col, 0)
        if margin_col == "profit_margin":
            margin_val = margin_val * 100

        risk_val = row.get(risk_col, 0) if risk_col else None
        rev_share = row[rev_col] / total_rev * 100 if total_rev > 0 else 0
        prod_count = int(row["product_id"])

        # Build intelligent commentary
        parts = [f"<b>{row['category']}</b>"]
        parts.append(f"contributes {rev_share:.1f}% of portfolio revenue")

        if margin_val > 35:
            parts.append(f"with <b>strong margins</b> ({margin_val:.1f}%)")
        elif margin_val > 20:
            parts.append(f"with <b>healthy margins</b> ({margin_val:.1f}%)")
        elif margin_val > 10:
            parts.append(f"with <b>moderate margins</b> ({margin_val:.1f}%)")
        else:
            parts.append(f"with <b>low margins</b> ({margin_val:.1f}%) — review needed")

        if risk_val is not None:
            if risk_val > 50:
                parts.append(f"and <b>elevated risk</b> ({risk_val:.0f}). Requires attention.")
            elif risk_val > 30:
                parts.append(f"and <b>moderate risk</b> ({risk_val:.0f}). Monitor regularly.")
            else:
                parts.append(f"and <b>low risk</b> ({risk_val:.0f}). Stable category.")

        commentary = " ".join(parts)
        story.append(Paragraph(f"• {commentary}", styles["CategoryCommentary"]))


def _build_pdf_risks_and_attention(story: List[Any], styles: Dict[str, ParagraphStyle],
                                    df: pd.DataFrame, insights: Dict[str, Any]) -> None:
    """
    Build the Top Risks & Attention Areas section.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
        insights: KPI insights dict.
    """
    story.append(Paragraph("7. Top Risks & Attention Areas", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_DANGER),
        spaceAfter=10, spaceBefore=0,
    ))

    risk_col = "composite_risk_score" if "composite_risk_score" in df.columns else None
    risk_level_col = "risk_level" if "risk_level" in df.columns else "risk_category"

    # Key risk observations
    story.append(Paragraph("Key Risk Observations", styles["SubSectionHeader"]))

    observations = []

    # 1. Categories with elevated risk
    if risk_col and "category" in df.columns:
        cat_risk = df.groupby("category")[risk_col].mean().sort_values(ascending=False)
        if not cat_risk.empty and cat_risk.iloc[0] > 50:
            observations.append(
                f"<b>High-Risk Category:</b> '{cat_risk.index[0]}' has the highest average "
                f"risk score ({cat_risk.iloc[0]:.1f}), significantly above "
                f"the portfolio average of {insights.get('avg_risk', 0):.1f}."
            )

    # 2. Products requiring review (top 3 high-risk)
    if risk_col:
        high_risk_products = df[df[risk_col] >= 50].sort_values(risk_col, ascending=False).head(3)
        if not high_risk_products.empty:
            prod_names = ", ".join(
                str(p) for p in high_risk_products["product_name"].head(3)
            )
            observations.append(
                f"<b>Products Requiring Review:</b> {prod_names} exhibit elevated "
                f"risk and should be prioritized for review."
            )

    # 3. Margin concerns
    margin_col = None
    for c in ["margin_percentage", "profit_margin"]:
        if c in df.columns:
            margin_col = c
            break
    if margin_col:
        low_margin_count = 0
        if margin_col == "profit_margin":
            low_margin_count = (df[margin_col] < 0.10).sum()
        else:
            low_margin_count = (df[margin_col] < 10).sum()
        if low_margin_count > 0:
            pct_low = low_margin_count / len(df) * 100
            observations.append(
                f"<b>Margin Concerns:</b> {low_margin_count} products ({pct_low:.1f}%) "
                f"have low margins (<10%), requiring margin improvement strategies."
            )

    # 4. Pricing anomalies (products with price changes > 30%)
    if "price_change_pct" in df.columns:
        high_change = df[abs(df["price_change_pct"]) > 30].shape[0]
        if high_change > 0:
            observations.append(
                f"<b>Pricing Anomalies:</b> {high_change} products show price change "
                f"recommendations exceeding 30%, suggesting potential data anomalies "
                f"or extreme market conditions."
            )

    # 5. Demand concerns
    if "demand_trend" in df.columns:
        declining = df[df["demand_trend"] < 0.4].shape[0]
        if declining > 0:
            observations.append(
                f"<b>Demand Decline:</b> {declining} products show weak demand signals "
                f"(demand trend < 0.4), requiring inventory and pricing strategy adjustments."
            )

    if not observations:
        observations.append(
            "<b>No critical risk flags identified.</b> Portfolio risk profile "
            "is within acceptable parameters."
        )

    for obs in observations:
        story.append(Paragraph(f"• {obs}", styles["BulletText"]))

    story.append(Spacer(1, 0.1 * inch))

    # High-risk products table
    if risk_col:
        high_risk = df[df[risk_col] >= 50].sort_values(risk_col, ascending=False).head(8)
        if not high_risk.empty:
            story.append(Paragraph("High-Risk Products Detail", styles["SubSectionHeader"]))
            hr_headers = ["Product", "Category", "Risk Score", "Level", "Primary Factor"]
            hr_data = [hr_headers]
            for _, row in high_risk.iterrows():
                hr_data.append([
                    str(row.get("product_name", "Unknown"))[:22],
                    str(row.get("category", ""))[:15],
                    str(round(row.get(risk_col, 0), 1)),
                    str(row.get(risk_level_col, ""))[:8],
                    str(row.get("primary_risk_factor", ""))[:25],
                ])

            hr_table = Table(hr_data, colWidths=[1.4 * inch, 1.0 * inch, 0.7 * inch, 0.7 * inch, 1.5 * inch])
            hr_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_DANGER)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0e0")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#fff5f5")]),
            ]))
            story.append(hr_table)


def _build_pdf_recommendations(story: List[Any], styles: Dict[str, ParagraphStyle],
                                df: pd.DataFrame, insights: Dict[str, Any]) -> None:
    """
    Build the Actionable Recommendations section.

    Args:
        story: ReportLab story list.
        styles: Paragraph style dict.
        df: Analysis DataFrame.
        insights: KPI insights dict.
    """
    story.append(Paragraph("8. Actionable Recommendations", styles["SectionHeader"]))
    story.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor(BRAND_ACCENT),
        spaceAfter=10, spaceBefore=0,
    ))

    story.append(Paragraph(
        "Based on the analysis conducted, the following business recommendations "
        "are provided for management consideration:",
        styles["BodyText2"],
    ))
    story.append(Spacer(1, 0.05 * inch))

    recommendations = []

    # 1. Price increases
    if insights.get("rec_increase", 0) > 0:
        recommendations.append({
            "action": "Implement Price Increases for Selected Products",
            "detail": f"{insights['rec_increase']} products identified for price optimization. "
                      f"Estimated revenue lift of ${insights.get('revenue_lift', 0):,.0f} "
                      f"and profit lift of ${insights.get('profit_lift', 0):,.0f}.",
            "priority": "High",
            "color": BRAND_ACCENT,
        })

    # 2. Monitor declining-demand products
    if "demand_trend" in df.columns:
        declining = df[df["demand_trend"] < 0.5].shape[0]
        if declining > 0:
            recommendations.append({
                "action": "Monitor & Adjust Declining-Demand Products",
                "detail": f"{declining} products show weakening demand signals. "
                          f"Consider promotional pricing, bundling, or inventory reduction strategies.",
                "priority": "Medium",
                "color": BRAND_WARNING,
            })

    # 3. Low margin categories
    margin_col = None
    for c in ["margin_percentage", "profit_margin"]:
        if c in df.columns:
            margin_col = c
            break
    if margin_col and "category" in df.columns:
        if margin_col == "profit_margin":
            low_margin_cats = df.groupby("category")[margin_col].mean()
            low_margin_cats = low_margin_cats[low_margin_cats < 0.15]
        else:
            low_margin_cats = df.groupby("category")[margin_col].mean()
            low_margin_cats = low_margin_cats[low_margin_cats < 15]
        if not low_margin_cats.empty:
            cat_names = ", ".join(low_margin_cats.index.tolist())
            recommendations.append({
                "action": "Review Low-Margin Categories",
                "detail": f"Categories ({cat_names}) have below-target margins. "
                          f"Review cost structures and pricing strategies.",
                "priority": "Medium",
                "color": BRAND_WARNING,
            })

    # 4. Inventory optimization
    if "inventory_level" in df.columns and "sales_volume" in df.columns:
        overstocked = df[df["inventory_level"] > df["sales_volume"] * 3].shape[0]
        if overstocked > 0:
            recommendations.append({
                "action": "Optimize Inventory Planning",
                "detail": f"{overstocked} products show elevated inventory relative to sales velocity. "
                          f"Review reorder points and inventory turnover strategies.",
                "priority": "Medium",
                "color": BRAND_SECONDARY,
            })

    # 5. Risk monitoring
    if insights.get("high_risk_count", 0) > 0:
        recommendations.append({
            "action": "Track High-Risk Products Closely",
            "detail": f"{insights['high_risk_count']} high-risk products require regular monitoring. "
                      f"Implement weekly review cycle for these items.",
            "priority": "High",
            "color": BRAND_DANGER,
        })

    # 6. Competitor-sensitive products
    if "competitor_price" in df.columns:
        competitive = df[abs(df["current_price"] - df["competitor_price"]) / df["competitor_price"] > 0.20].shape[0]
        if competitive > 0:
            recommendations.append({
                "action": "Track Competitor-Sensitive Products",
                "detail": f"{competitive} products have prices diverging significantly "
                          f"(>20%) from competitor benchmarks. Monitor competitive positioning.",
                "priority": "Low",
                "color": BRAND_TEAL,
            })

    # Add revenue concentration recommendation
    rev_col = "expected_revenue" if "expected_revenue" in df.columns else "revenue"
    if rev_col in df.columns:
        top10_share = df.nlargest(10, rev_col)[rev_col].sum() / insights.get("total_revenue", 1) * 100
        if top10_share > 50:
            recommendations.append({
                "action": "Diversify Revenue Concentration",
                "detail": f"Top 10 products contribute {top10_share:.0f}% of revenue. "
                          f"Explore growth strategies for mid-tier products to reduce dependency.",
                "priority": "Medium",
                "color": BRAND_WARNING,
            })

    if not recommendations:
        recommendations.append({
            "action": "Portfolio is Well-Positioned",
            "detail": "No critical recommendations identified. Continue regular monitoring "
                      "and periodic pricing reviews.",
            "priority": "Info",
            "color": BRAND_ACCENT,
        })

    # Render recommendations as styled cards
    for i, rec in enumerate(recommendations):
        priority_badge = f"[{rec['priority']}]" if rec.get("priority") else ""

        # Build recommendation card using a table
        rec_cell = Paragraph(
            f"<b>{rec['action']}</b> {priority_badge}<br/>"
            f"<font size='8' color='#546e7a'>{rec['detail']}</font>",
            styles["Normal"],
        )

        rec_table = Table(
            [[rec_cell]],
            colWidths=[6.3 * inch],
        )
        rec_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#e0e0e0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fafafa")),
            ("LINELEFT", (0, 0), (0, 0), 4, colors.HexColor(rec.get("color", BRAND_PRIMARY))),
        ]))
        story.append(rec_table)
        story.append(Spacer(1, 0.06 * inch))


def _build_pdf_footer(canvas, doc):
    """Add page numbers and footer to each page."""
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#9e9e9e"))
    canvas.drawCentredString(
        PAGE_WIDTH / 2, 0.4 * inch,
        f"AI Pricing Intelligence Platform  |  Page {doc.page}"
    )
    canvas.drawCentredString(
        PAGE_WIDTH / 2, 0.25 * inch,
        "CONFIDENTIAL — For Internal Use Only"
    )
    canvas.restoreState()


# Global placeholder for df reference in narratives (used by _build_pdf_executive_summary)
df_global_placeholder = None


def generate_pdf_report(
    df: pd.DataFrame,
    insights: Optional[Dict[str, Any]] = None,
    filename: Optional[str] = None,
) -> str:
    """
    Generate a comprehensive, professionally styled PDF report.

    Sections:
    1. Cover Page — branding, metadata, quick KPI snapshot
    2. Executive Summary — narrative business summary
    3. Executive KPI Dashboard — KPI cards with explanations
    4. Business Insights — pricing, revenue, category insights
    5. Pricing Action Plan — increase/reduce/maintain groups
    6. Visual Analytics — pie, bar, donut charts
    7. Category Performance — table + business commentary
    8. Top Risks & Attention Areas — risk observations + high-risk table
    9. Actionable Recommendations — business next steps

    Args:
        df: Analysis DataFrame with all computed columns.
        insights: Dict of KPI metrics (auto-computed if None).
        filename: Output filename (auto-generated if None).

    Returns:
        Path to the generated PDF file.
    """
    global df_global_placeholder
    df_global_placeholder = df

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if filename is None:
        filename = f"pricing_report_{timestamp}.pdf"

    config = AppConfig()
    filepath = config.get_report_path(filename)

    styles = _get_pdf_styles()
    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=letter,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        title="Executive Pricing Analysis Report",
        author="AI Pricing Intelligence Platform",
    )

    if insights is None:
        insights = _compute_insights(df)

    # Generate narrative insights
    narratives = _generate_narrative_insights(df, insights)

    story: List[Any] = []

    # =========================================================
    # PAGE 1: Cover Page
    # =========================================================
    _build_pdf_cover_page(story, styles, insights)
    doc.build(story, onFirstPage=_build_pdf_footer, onLaterPages=_build_pdf_footer)
    story = []

    # Clear placeholder
    df_global_placeholder = None

    # =========================================================
    # PAGE 2: Executive Summary + KPI Dashboard
    # =========================================================
    _build_pdf_executive_summary(story, styles, insights, narratives)
    story.append(Spacer(1, 0.15 * inch))
    _build_pdf_kpi_dashboard(story, styles, insights)

    # =========================================================
    # PAGE 3+: Business Insights + Pricing Action Plan
    # =========================================================
    story.append(PageBreak())
    _build_pdf_business_insights(story, styles, df, insights, narratives)
    story.append(Spacer(1, 0.15 * inch))
    _build_pdf_pricing_action_plan(story, styles, df)

    # =========================================================
    # PAGE 4+: Visualizations
    # =========================================================
    story.append(PageBreak())
    _build_pdf_visualizations(story, styles, df)

    # =========================================================
    # PAGE 5+: Category Performance + Risks
    # =========================================================
    story.append(PageBreak())
    _build_pdf_category_performance(story, styles, df, insights)
    story.append(Spacer(1, 0.15 * inch))
    _build_pdf_risks_and_attention(story, styles, df, insights)

    # =========================================================
    # PAGE 6+: Actionable Recommendations
    # =========================================================
    story.append(PageBreak())
    _build_pdf_recommendations(story, styles, df, insights)

    # Build the document with footer on all pages
    doc.build(story, onFirstPage=_build_pdf_footer, onLaterPages=_build_pdf_footer)

    logger.info(
        f"PDF report generated: {filepath} "
        f"({os.path.getsize(filepath):,} bytes)"
    )
    return str(filepath)


# ═══════════════════════════════════════════════════════════════════════════
# CLASS-BASED API  (backward-compatible)
# ═══════════════════════════════════════════════════════════════════════════

class ReportGenerator:
    """
    Enterprise report generator for the AI Pricing Intelligence Platform.

    Generates reports in three formats:
    - CSV  (portable data)
    - Excel (multi-sheet workbook with charts, OpenPyXL)
    - PDF  (styled document, ReportLab)

    Each report includes KPIs, portfolio summary, risk analysis,
    pricing recommendations, and top product tables.

    Usage:
        generator = ReportGenerator()
        path = generator.generate_pdf_report(df)
        path = generator.generate_excel_report(df)
        path = generator.generate_csv_report(df)
    """

    def __init__(self, config: Optional[AppConfig] = None) -> None:
        """
        Initialize the report generator.

        Args:
            config: Application configuration.
        """
        self.config = config or AppConfig()

    def generate_csv_report(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
        **kwargs: Any,
    ) -> str:
        """
        Generate a CSV report from the analysis DataFrame.

        Args:
            df: Analysis DataFrame.
            filename: Output filename.
            **kwargs: Additional CSV args.

        Returns:
            Path to generated CSV file.
        """
        return generate_csv_report(df, filename=filename, **kwargs)

    def generate_excel_report(
        self,
        df: pd.DataFrame,
        insights: Optional[Dict[str, Any]] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        Generate a formatted multi-sheet Excel report.

        Args:
            df: Analysis DataFrame.
            insights: KPI metrics dict.
            filename: Output filename.

        Returns:
            Path to generated Excel file.
        """
        return generate_excel_report(df, insights=insights, filename=filename)

    def generate_pdf_report(
        self,
        df: pd.DataFrame,
        insights: Optional[Dict[str, Any]] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        Generate a professionally styled PDF report.

        Args:
            df: Analysis DataFrame.
            insights: KPI metrics dict.
            filename: Output filename.

        Returns:
            Path to generated PDF file.
        """
        return generate_pdf_report(df, insights=insights, filename=filename)

    def generate_risk_report(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
    ) -> str:
        """
        Generate a focused risk assessment Excel report.

        Args:
            df: DataFrame with risk analysis columns.
            filename: Output filename.

        Returns:
            Path to generated Excel file.
        """
        risk_cols = [
            "product_name", "category",
            "composite_risk_score", "risk_level",
            "primary_risk_factor",
            "demand_risk_score", "profitability_risk_score",
            "inventory_risk_score", "competitor_risk_score",
            "margin_risk_score",
        ]
        available = [c for c in risk_cols if c in df.columns]

        risk_df = df[available].sort_values(
            "composite_risk_score", ascending=False
        )

        insights = _compute_insights(risk_df)
        insights["report_type"] = "Risk Assessment"
        insights["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"risk_report_{timestamp}.xlsx"

        return self.generate_excel_report(
            risk_df, insights=insights, filename=filename
        )